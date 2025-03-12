import openpyxl

# Crear un nuevo libro de Excel
wb = openpyxl.Workbook()
ws = wb.active

# Encabezados
ws.append(["Iteración", "xₙ", "f(xₙ)", "f'(xₙ)", "xₙ₊₁", "Error"])

# Valor inicial
x_n = 5.0
ws["B2"] = x_n

# Fórmulas para la primera iteración
ws["C2"] = f"=B2^6 - 273"
ws["D2"] = f"=6*B2^5"
ws["E2"] = f"=B2 - C2/D2"
ws["F2"] = f"=ABS(E2 - B2)"

# Iteraciones subsiguientes
for i in range(1, 11):
    current_row = str(2 + i)
    prev_row = str(1 + i)
    ws[f"A{current_row}"] = i
    ws[f"B{current_row}"] = f"=E{prev_row}"  # xₙ = xₙ₊₁ anterior
    ws[f"C{current_row}"] = f"=B{current_row}^6 - 273"
    ws[f"D{current_row}"] = f"=6*B{current_row}^5"
    ws[f"E{current_row}"] = f"=B{current_row} - C{current_row}/D{current_row}"
    ws[f"F{current_row}"] = f"=ABS(E{current_row} - B{current_row})"

# Guardar el archivo
wb.save("newton_raphson.xlsx")